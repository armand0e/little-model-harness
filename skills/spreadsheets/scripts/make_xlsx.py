"""Build a polished .xlsx from a JSON spec.

Usage: python make_xlsx.py spec.json output.xlsx
Spec: {"sheets": [{
  "name", "headers", "rows",
  "title"?: str,               # bold title block above the table
  "number_formats"?: {"B": "#,##0.00"},
  "widths"?: {"A": 18},
  "totals"?: true,             # SUM row for numeric columns
  "chart"?: {"type": "bar"|"line"|"pie", "labels": "A",
              "values": ["B", "C"], "title"?: str}
}]}
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ACCENT = "1F4E79"
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=15, color="262626")
TOTAL_FONT = Font(bold=True, color="262626")
TOTAL_BORDER = Border(top=Side(style="medium", color=ACCENT))
MAX_SHEETS = 100
MAX_ROWS = 100_000
MAX_COLUMNS = 500
XML_INVALID_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _sheet_name(value, used):
    name = re.sub(r"[\\/*?:\[\]]", "_", str(value or "Sheet")).strip("'")[:31]
    name = name or "Sheet"
    base = name
    suffix = 2
    while name.casefold() in used:
        marker = f" ({suffix})"
        name = base[:31 - len(marker)] + marker
        suffix += 1
    used.add(name.casefold())
    return name


def _clean_cell(value):
    return XML_INVALID_RE.sub("", value) if isinstance(value, str) else value


def build(spec_path, out_path):
    with open(spec_path, encoding="utf-8") as f:
        spec = json.load(f)
    if not isinstance(spec, dict) or not isinstance(spec.get("sheets", []), list):
        sys.exit("Error: spec must be an object with a 'sheets' array")
    if len(spec.get("sheets", [])) > MAX_SHEETS:
        sys.exit(f"Error: at most {MAX_SHEETS} sheets are supported")
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()
    for sheet in spec.get("sheets", []):
        if not isinstance(sheet, dict):
            sys.exit("Error: each sheet must be an object")
        ws = wb.create_sheet(title=_sheet_name(sheet.get("name"), used_names))
        headers = sheet.get("headers") or []
        rows = sheet.get("rows") or []
        if not isinstance(headers, list) or not isinstance(rows, list) \
                or not all(isinstance(row, list) for row in rows):
            sys.exit("Error: sheet headers and rows must be arrays")
        if len(rows) > MAX_ROWS:
            sys.exit(f"Error: a sheet may contain at most {MAX_ROWS:,} rows")
        if max([len(headers), *(len(row) for row in rows)], default=0) > MAX_COLUMNS:
            sys.exit(f"Error: a sheet may contain at most {MAX_COLUMNS} columns")
        if any(isinstance(value, (dict, list))
               for row in [headers, *rows] for value in row):
            sys.exit("Error: spreadsheet cells must be scalar JSON values")

        # optional title block above the data
        header_row = 1
        title = sheet.get("title")
        if title:
            ws.append([_clean_cell(str(title))])
            ws.cell(row=1, column=1).font = TITLE_FONT
            ws.append([])
            header_row = 3
        if headers:
            ws.append([_clean_cell(value) for value in headers])
            for cell in ws[header_row]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")
            ws.row_dimensions[header_row].height = 20
            ws.freeze_panes = f"A{header_row + 1}"
        for row in rows:
            ws.append([_clean_cell(value) for value in row])

        # totals row: SUM for numeric columns, label in the first column
        data_first = header_row + 1
        data_last = header_row + len(rows)
        numeric_columns = set()
        for column_index in range(1, len(headers) + 1):
            values = [ws.cell(row=r, column=column_index).value
                      for r in range(data_first, data_last + 1)]
            present = [v for v in values if v is not None and v != ""]
            if present and all(isinstance(v, (int, float)) for v in present):
                numeric_columns.add(column_index)
        if sheet.get("totals") and rows and headers:
            total_row = data_last + 1
            for column_index in range(1, len(headers) + 1):
                cell = ws.cell(row=total_row, column=column_index)
                if column_index == 1 and column_index not in numeric_columns:
                    cell.value = "Total"
                elif column_index in numeric_columns:
                    letter = get_column_letter(column_index)
                    cell.value = (f"=SUM({letter}{data_first}:"
                                  f"{letter}{data_last})")
                cell.font = TOTAL_FONT
                cell.border = TOTAL_BORDER

        # a real Excel table: filters + banded rows, in an accent style
        if headers and rows:
            ref = (f"A{header_row}:"
                   f"{get_column_letter(len(headers))}{data_last}")
            table = Table(displayName=f"Table{len(wb.sheetnames)}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True)
            try:
                ws.add_table(table)
            except ValueError:
                pass  # duplicate/invalid refs fall back to plain styling
        ws.sheet_view.showGridLines = False

        # number formats per column letter
        number_formats = sheet.get("number_formats") or {}
        widths = sheet.get("widths") or {}
        if not isinstance(number_formats, dict) or not isinstance(widths, dict):
            sys.exit("Error: number_formats and widths must be objects")
        for col, fmt in number_formats.items():
            if not re.fullmatch(r"[A-Za-z]{1,3}", str(col)) or not isinstance(fmt, str):
                sys.exit("Error: number_formats must map column letters to strings")
            for cell in ws[col]:
                if cell.row > header_row:
                    cell.number_format = fmt
        # column widths: explicit or auto-fit from content
        ncols = max([len(headers)] + [len(r) for r in rows] or [1])
        for idx in range(1, ncols + 1):
            letter = get_column_letter(idx)
            if letter in widths:
                try:
                    width = float(widths[letter])
                except (TypeError, ValueError):
                    sys.exit(f"Error: width for column {letter} must be numeric")
                ws.column_dimensions[letter].width = min(max(width, 1), 255)
            else:
                longest = max((len(str(c.value)) for c in ws[letter]
                               if c.value is not None), default=8)
                ws.column_dimensions[letter].width = min(max(longest + 2, 10), 45)

        # optional chart anchored beside the data
        chart_spec = sheet.get("chart")
        if isinstance(chart_spec, dict) and headers and rows:
            kinds = {"bar": BarChart, "line": LineChart, "pie": PieChart}
            chart_type = str(chart_spec.get("type", "bar")).lower()
            chart = kinds.get(chart_type, BarChart)()
            chart.title = _clean_cell(str(
                chart_spec.get("title") or title or "Chart"))
            chart.style = 10
            chart.height = 8.5
            chart.width = 16
            label_letter = str(chart_spec.get("labels", "A")).upper()
            if not re.fullmatch(r"[A-Z]{1,3}", label_letter):
                sys.exit("Error: chart labels must be a column letter")
            value_letters = chart_spec.get("values") or []
            if not isinstance(value_letters, list) or not all(
                    re.fullmatch(r"[A-Za-z]{1,3}", str(v))
                    for v in value_letters):
                sys.exit("Error: chart values must be column letters")
            from openpyxl.utils import column_index_from_string
            labels = Reference(
                ws, min_col=column_index_from_string(label_letter),
                min_row=data_first, max_row=data_last)
            for letter in value_letters[:6]:
                column_number = column_index_from_string(str(letter).upper())
                data = Reference(ws, min_col=column_number,
                                 min_row=header_row, max_row=data_last)
                chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)
            anchor_col = get_column_letter(len(headers) + 2)
            ws.add_chart(chart, f"{anchor_col}{header_row}")
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")
    output = Path(out_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Saved {out_path} ({len(wb.sheetnames)} sheet(s): {', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit("Usage: python make_xlsx.py spec.json output.xlsx")
    build(sys.argv[1], sys.argv[2])
