"""Dump .xlsx contents as text tables.

Usage: python read_xlsx.py file.xlsx [sheet_name] [max_rows]
"""
import sys

from openpyxl import load_workbook

MAX_CHARS = 8000
MAX_COLUMNS = 100


def main(path, sheet=None, max_rows=40):
    max_rows = min(max(int(max_rows), 1), 1000)
    wb = load_workbook(path, data_only=False, read_only=True)
    names = [sheet] if sheet else wb.sheetnames
    out = []
    for name in names:
        if name not in wb.sheetnames:
            out.append(f"(no sheet named {name}; sheets: {', '.join(wb.sheetnames)})")
            continue
        ws = wb[name]
        out.append(f"=== Sheet: {name} ({ws.max_row} rows x {ws.max_column} cols) ===")
        for r, row in enumerate(ws.iter_rows(
                min_col=1, max_col=min(ws.max_column, MAX_COLUMNS),
                values_only=True), 1):
            if r > max_rows:
                out.append(f"...[{ws.max_row - max_rows} more rows]")
                break
            out.append(" | ".join("" if v is None else str(v) for v in row))
            if sum(len(line) for line in out) > MAX_CHARS:
                out.append("...[output limit reached; request fewer rows]")
                break
        out.append("")
    text = "\n".join(out)
    if len(text) > MAX_CHARS:
        text = text[:MAX_CHARS] + "\n...[truncated; pass a sheet name / smaller max_rows]"
    print(text)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Usage: python read_xlsx.py file.xlsx [sheet] [max_rows]")
    main(sys.argv[1],
         sys.argv[2] if len(sys.argv) > 2 else None,
         int(sys.argv[3]) if len(sys.argv) > 3 else 40)
